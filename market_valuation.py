"""
市場全体のPER・PBR（月次）の取得と保存。

なぜ指数のPERではないのか:
  Yahooは指数(^N225 / ^GSPC)にバリュエーションを持たせておらず、日経が公表する
  日経平均PERは配信ページが機械取得を拒否している。一方、東証プライム全体のPERは
  JPXがExcelで、S&P500のPERは multpl.com が月次で公開している。
  「日本市場と米国市場、同じ利益にいくらの値段が付いているか」を見るのが目的なので、
  225銘柄に限定する必要はなく、市場全体の方がむしろ話として正しい。

⚠️ 日経平均PERとは別物。画面には必ず出所と定義を書くこと。

JPXのExcelの罠:
  2022年3月以前のファイルは、数値がキャッシュ無しの数式（"=17.2"）として
  入っている。openpyxl を data_only=True で読むと全部 None になる。
  data_only=False で読んで先頭の "=" を落とす。新しいファイルは素の数値なので
  どちらも同じ処理で通る。
"""

import io
import json
import os
import re
import time
from datetime import datetime, timezone

import requests

JPX_INDEX_URL = 'https://www.jpx.co.jp/markets/statistics-equities/misc/04.html'
JPX_BASE = 'https://www.jpx.co.jp'
JPX_SHEET = '規模別・業種別（連結）'

# 東証は2022年4月に市場区分を再編した。連続した1本の系列として扱う。
JP_MARKET_NAMES = ('プライム市場', '市場一部', '市場第一部')

MULTPL_PE_URL = 'https://www.multpl.com/s-p-500-pe-ratio/table/by-month'

UA = {'User-Agent': 'Mozilla/5.0 (compatible; CompanyNote/1.0)'}

MARKETS = {
    'jp_prime': {
        'name': '日本（東証プライム）',
        'short_name': '日本',
        'note': '東証プライム市場・総合の加重平均PER（連結）。2022年3月以前は東証一部。出所: JPX',
        'color': '#1b4332',
    },
    'sp500': {
        'name': '米国（S&P500）',
        'short_name': '米国',
        'note': 'S&P500の実績PER。出所: multpl.com',
        'color': '#d64545',
    },
}


# ---------------------------------------------------------------
# JPX（日本）
# ---------------------------------------------------------------

def _cell_number(value):
    """Excelのセルを数値にする。'=17.2' のような数式表記にも対応する。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lstrip('=').strip()
    if not text or text in ('-', '－', '＊', '*'):
        # JPXの凡例: '－'は該当値なしまたはマイナス、'＊'はPER1000倍以上
        return None
    try:
        return float(text)
    except ValueError:
        return None


def list_jpx_files():
    """JPXの一覧ページから {YYYYMM: URL} を作る。URLに不透明なIDが入るため毎回引く。"""
    res = requests.get(JPX_INDEX_URL, headers=UA, timeout=30)
    res.raise_for_status()
    found = re.findall(r'href="([^"]+perpbr(\d{6})\.xlsx)"', res.text)
    return {ym: (path if path.startswith('http') else JPX_BASE + path)
            for path, ym in found}


def parse_jpx_workbook(content):
    """月次Excelから東証プライム（旧・東証一部）総合の行を取り出す。

    列: 0年月 1市場区分名 2Section 3種別 4Industry 5会社数
        6単純PER 7単純PBR 8EPS 9BPS 10加重PER 11加重PBR
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    if JPX_SHEET not in wb.sheetnames:
        return None
    ws = wb[JPX_SHEET]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=12, values_only=True):
        market = (str(row[1]).strip() if row[1] else '')
        kind = (str(row[3]).strip() if row[3] else '')
        if market not in JP_MARKET_NAMES or kind != '総合':
            continue
        return {
            'company_count': int(_cell_number(row[5]) or 0) or None,
            'per_simple': _cell_number(row[6]),
            'pbr_simple': _cell_number(row[7]),
            'per': _cell_number(row[10]),
            'pbr': _cell_number(row[11]),
        }
    return None


def fetch_jpx(months=None, sleep=0.4, progress=None):
    """JPXの月次Excelを取得して [{month, per, ...}] を返す。

    Args:
        months: 取得する 'YYYYMM' の集合。Noneなら公開されている全月。
        progress: (done, total) を受け取るコールバック
    """
    files = list_jpx_files()
    targets = sorted(ym for ym in files if months is None or ym in months)

    rows = []
    for i, ym in enumerate(targets):
        try:
            res = requests.get(files[ym], headers=UA, timeout=45)
            res.raise_for_status()
            parsed = parse_jpx_workbook(res.content)
        except Exception as e:
            print(f'JPX {ym} の取得エラー: {e}')
            parsed = None

        if parsed and (parsed.get('per') or parsed.get('per_simple')):
            rows.append({
                'market_key': 'jp_prime',
                'month': f'{ym[:4]}-{ym[4:]}-01',
                'source': 'JPX 規模別・業種別PER・PBR（連結）',
                **parsed,
            })

        if progress:
            progress(i + 1, len(targets))
        if sleep:
            time.sleep(sleep)

    return rows


# ---------------------------------------------------------------
# multpl（米国）
# ---------------------------------------------------------------

def fetch_sp500():
    """multpl.com の月次PERを [{month, per}] で返す。

    表は新しい順で、当月だけ「Jul 27, 2026」のような最新日の行が
    月初の行より前に来る。月ごとに最初に現れた行＝新しい方を採用する。
    """
    import pandas as pd

    res = requests.get(MULTPL_PE_URL, headers=UA, timeout=30)
    res.raise_for_status()
    table = pd.read_html(io.StringIO(res.text))[0]

    seen = set()
    rows = []
    for _, r in table.iterrows():
        try:
            dt = pd.to_datetime(str(r.iloc[0]))
        except Exception:
            continue
        key = (dt.year, dt.month)
        if key in seen:
            continue

        # 推計値を示す記号が混ざるので数字だけ拾う
        m = re.search(r'-?\d+(?:\.\d+)?', str(r.iloc[1]))
        if not m:
            continue

        seen.add(key)
        rows.append({
            'market_key': 'sp500',
            'month': f'{dt.year:04d}-{dt.month:02d}-01',
            'per': float(m.group()),
            'source': 'multpl.com S&P 500 PE Ratio',
        })

    rows.sort(key=lambda x: x['month'])
    return rows


# ---------------------------------------------------------------
# 保存（同梱JSON＋起動後の差分取得）
#
# DBに置かない理由:
#   月1行しか増えない、全部で200行程度のデータのためにテーブルを増やすと
#   マイグレーションの適用待ちが発生して、その間ページが動かない。
#   履歴はリポジトリ同梱のJSONに固め、足りない月だけ起動後に取りに行く。
#   （JPXの過去の数値は確定値で、後から書き換わらないので同梱で問題ない）
# ---------------------------------------------------------------

SNAPSHOT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             'static', 'market_valuation.json')

# 取得しに行った結果のキャッシュ。JPXは月次更新なので長めに持つ。
_CACHE_TTL = 12 * 3600
_cache = {'at': 0, 'series': None}


def load_snapshot():
    """同梱JSONを読む。無ければ空を返す。"""
    try:
        with io.open(SNAPSHOT_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'series': {}, 'generated_at': None}


def _month_key(month_str):
    return month_str[:4] + month_str[5:7]


def _missing_months(series):
    """同梱JSONの最終月から今月までの 'YYYYMM' を返す（JPX用）。"""
    from datetime import date

    rows = series.get('jp_prime') or []
    if not rows:
        return None  # 全期間取り直し

    last = _month_key(rows[-1]['month'])
    today = date.today()
    out = []
    y, m = int(last[:4]), int(last[4:])
    while True:
        m += 1
        if m == 13:
            y, m = y + 1, 1
        if (y, m) > (today.year, today.month):
            break
        out.append(f'{y:04d}{m:02d}')
    return out


def _merge(series, rows):
    """月をキーに上書きマージして昇順に並べ直す"""
    for r in rows:
        key = r['market_key']
        bucket = {x['month']: x for x in series.get(key, [])}
        bucket[r['month']] = {k: v for k, v in r.items() if k != 'market_key'}
        series[key] = sorted(bucket.values(), key=lambda x: x['month'])
    return series


def get_series(force=False):
    """月次系列を {market_key: [{month, per, ...}]} で返す。

    同梱JSONを土台に、公開済みで手元に無い月だけ取りに行く。
    通常は0〜3ファイルで済む。取得に失敗しても同梱ぶんは必ず返る。
    """
    now = time.time()
    if not force and _cache['series'] and now - _cache['at'] < _CACHE_TTL:
        return _cache['series']

    snapshot = load_snapshot()
    series = {k: list(v) for k, v in (snapshot.get('series') or {}).items()}

    # 日本: 足りない月だけ
    try:
        missing = _missing_months(series)
        if missing is None or missing:
            series = _merge(series, fetch_jpx(months=(set(missing) if missing else None)))
    except Exception as e:
        print(f'JPXの差分取得に失敗（同梱ぶんで継続）: {e}')

    # 米国: 表が1枚なので毎回まとめて取り直す（1リクエスト）。
    # 同梱JSONと同じ範囲・同じ形に揃えてから差し替える。
    try:
        us = fetch_sp500()
        if us:
            series['sp500'] = [
                {k: v for k, v in r.items() if k not in ('market_key', 'source')}
                for r in us if r['month'] >= '1990-01-01'
            ]
    except Exception as e:
        print(f'S&P500 PERの取得に失敗（同梱ぶんで継続）: {e}')

    _cache['at'] = now
    _cache['series'] = series
    return series


def rebuild_snapshot(progress=None):
    """公開されている全期間を取り直して同梱JSONを書き出す。

    リポジトリに commit するためのもの。通常運転では呼ばない。
        python market_valuation.py --rebuild
    """
    series = {}
    series = _merge(series, fetch_jpx(progress=progress))
    series = _merge(series, fetch_sp500())

    # 出所は行ごとに持たず MARKETS 側に1つ持つ（同じ文字列を全行に置くと数倍になる）。
    # S&P500は1871年まで遡れるが、比較相手の日本が2020年からなので1990年で切る。
    for key, rows in series.items():
        rows = [{k: v for k, v in r.items() if k != 'source'} for r in rows]
        if key == 'sp500':
            rows = [r for r in rows if r['month'] >= '1990-01-01']
        series[key] = rows

    payload = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'markets': MARKETS,
        'series': series,
    }
    os.makedirs(os.path.dirname(SNAPSHOT_PATH), exist_ok=True)
    with io.open(SNAPSHOT_PATH, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    return {k: len(v) for k, v in series.items()}


if __name__ == '__main__':
    import sys

    if '--rebuild' in sys.argv:
        def _p(done, total):
            if done % 10 == 0 or done == total:
                print(f'  JPX {done}/{total}')
        print('全期間を取得しています…')
        print('完了:', rebuild_snapshot(progress=_p))
        print('出力:', SNAPSHOT_PATH)
    else:
        print('使い方: python market_valuation.py --rebuild')
